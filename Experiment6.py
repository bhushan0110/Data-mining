import openpyxl
import math
from itertools import combinations


def read_data_in_dict(path):
  wb=openpyxl.load_workbook(path) 
  sheet_obj = wb.active
  row = sheet_obj.max_row
  col = sheet_obj.max_column

  trans = []
  for i in range(2,row+1):
    x = sheet_obj.cell(row=i,column=1)
    trans.append(x.value)

  string_data=[]
  for i in range(2,row+1):
    for j in range(2,col+1):
      x = sheet_obj.cell(row=i,column=j)
      if(x.value):
        string_data.append(x.value)
  items = set(string_data)
  items = sorted(items)

  length = len(items)
  transactions=[]
  for i in range(2,row+1):
    s=[]
    for j in range(2,col+1):
      x = sheet_obj.cell(row=i,column=j)
      if(x.value):
        s.append(x.value)
    s=sorted(s)  
    req=[]
    k=0
    for j in range(0,length):
      if k<len(s):
        if items[j]==s[k]:
          req.append(1)
          k=k+1
        else:
          req.append(0)
      else:
        req.append(0)

    transactions.append(req)

  data={
      'items':items,
      'transactions': transactions
  }
  return data

data = read_data_in_dict("/content/Experiment6.xlsx")


# In[4]:

def get_freq(s,items,transactions):
    freq=0
    for t in transactions:
        temp=1
        for item in s:
            temp*=t[items.index(item)]
        if temp==1:
            freq+=1  
    return freq

def frequent_itemsets(data,level,min_support):
    items = data['items']
    transactions = data['transactions']
    min_freq = math.ceil(min_support*len(transactions))
    sets = list(combinations(items,level))
    frequent_sets = []
    for s in sets:
        freq=get_freq(s,items,transactions)
        if freq>=min_freq:
            frequent_sets.append(s)
    return frequent_sets
        
x=len(data['items'])
min_sup=float(input("Enter Min Support: "))


print("Min Support: ",min_sup,"\n")
for i in range(1,x):
  arr=frequent_itemsets(data,i,min_sup)
  if(len(arr)>0):
    print("Level:",i," ",arr)
