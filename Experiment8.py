import openpyxl
import math

wb=openpyxl.load_workbook("/content/Experiment8.xlsx") 
sheet_obj = wb.active
row = sheet_obj.max_row
col = sheet_obj.max_column

def std_dev(X,l,x_mean):
  ans=0
  for i in range(0,l):
    ans = ans + (X[i]-x_mean)*(X[i]-x_mean)
  return ans

X=[]
Y=[]
x_mean=0
y_mean=0
for i in range(2,row+1):
  val=sheet_obj.cell(row=i,column=2)
  val2=sheet_obj.cell(row=i,column=3)
  X.append(val.value)
  x_mean=x_mean+val.value
  Y.append(val2.value)
  y_mean=y_mean+val2.value

l=len(X)

x_mean=x_mean/l
y_mean=y_mean/l

cov=0
for i in range(0,l):
  cov=cov + (X[i]-x_mean)*(Y[i]-y_mean)


s = math.sqrt(std_dev(X,l,x_mean)*std_dev(Y,l,y_mean))

corelation = cov/s
print(corelation)